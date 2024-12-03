import pandas as pd
import numpy as np
import logging
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

# Configuración del logger
logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")
logger = logging.getLogger(__name__)

def ajustar_ancho_columnas(archivo):
    # Cargar el archivo Excel con openpyxl
    wb = openpyxl.load_workbook(archivo)
    hoja = wb.active

    # Ajustar automáticamente el ancho de cada columna según el contenido más largo
    for columna in hoja.columns:
        max_length = 0
        columna_letra = columna[0].column_letter  # Obtener la letra de la columna
        for celda in columna:
            try:
                if celda.value:
                    max_length = max(max_length, len(str(celda.value)))
            except Exception as e:
                logger.warning(f"Error ajustando la celda {celda.coordinate}: {e}")

        # Asignar ancho ajustado a la columna
        hoja.column_dimensions[columna_letra].width = max_length + 2  
        logger.info("Aplicando formato a los encabezados...")
    encabezado_fila = hoja[1]  # Primera fila (encabezados)
    estilo_encabezado = {
        "fill": PatternFill(start_color="C6D8FD", end_color="C6D8FD", fill_type="solid"),
        "font": Font(size=12, bold=True, color="000000"),
        "alignment": Alignment(horizontal="center", vertical="center"),
    }

    for celda in encabezado_fila:
        celda.fill = estilo_encabezado["fill"]
        celda.font = estilo_encabezado["font"]
        celda.alignment = estilo_encabezado["alignment"]    
            # Margen adicional para mayor claridad

    # Guardar los cambios en el archivo
    wb.save(archivo)
    logger.info(f"Ancho de columnas ajustado automáticamente en el archivo: {archivo}")
def procesar_archivo_instructores(file_instru, file_sofia):
    # Leer archivo de instructores
    logger.info("Leyendo el archivo de instructores...")
    instru_df = pd.read_excel(file_instru, sheet_name="Validar nombre aprendiz", header=None, dtype=str)
    logger.info("Archivo de instructores leído correctamente.")

    # Validar que la celda D2 contiene el Código de Ficha
    logger.info("Extrayendo el Código de Ficha de la celda D2...")
    try:
        codigo_ficha = instru_df.iloc[1, 3]  # Fila 2, columna 4 (D2)
        if pd.isna(codigo_ficha):
            raise ValueError("La celda D2 está vacía o no contiene un Código de Ficha válido.")
        logger.info(f"Código de Ficha identificado: {codigo_ficha}")
    except IndexError:
        logger.error("No se encontró la celda D2 en la hoja especificada.")
        raise ValueError("No se encontró la celda D2 en la hoja especificada.")

    # Tomar encabezados de la fila 10
    logger.info("Extrayendo encabezados de la fila 10...")
    encabezados = instru_df.iloc[9]
    instru_df = instru_df.iloc[10:]  # Eliminar filas hasta la fila 10
    instru_df.columns = encabezados
    instru_df.reset_index(drop=True, inplace=True)
    logger.info(f"Columnas extraídas del archivo de instructores: {list(instru_df.columns)}")

    # Renombrar columnas de instructores
    logger.info("Renombrando columnas en el archivo de instructores...")
    instru_df.rename(columns={
        "Tipo": "TIPO_INSTRU",
        "Documento de identificación": "DOCUMENTO_DE_IDENTIFICACION_INSTRU",
        "Nombre completo SENA": "NOMBRE_COMPLETO_SENA_INSTRU"
    }, inplace=True)
    logger.info(f"Columnas renombradas de instructores: {list(instru_df.columns)}")

    # Filtrar datos sin nombre o identificación
    logger.info("Filtrando filas sin nombre o identificación en instructores...")
    instru_df.dropna(subset=["NOMBRE_COMPLETO_SENA_INSTRU", "DOCUMENTO_DE_IDENTIFICACION_INSTRU"], inplace=True)

    # Leer archivo Sofía desde la fila correcta
    logger.info("Leyendo el archivo de Sofía...")
    sofia_df = pd.read_excel(file_sofia, header=1)  # Leer desde fila 2
    sofia_df.rename(columns={
        "Ficha": "FICHA",
        "Tipo Programa": "TIPO_PROGRAMA",
        "Nivel Formación": "NIVEL_FORMACION",
        "Denominación Programa": "DENOMINACION_PROGRAMA",
        "Tipo Documento": "TIPO_SOFIA",
        "No. Documento": "DOCUMENTO_DE_IDENTIFICACION_SOFIA",
        "Nombre Aprendiz": "NOMBRE_COMPLETO_SENA_SOFIA"
    }, inplace=True)
    logger.info(f"Columnas extraídas del archivo de Sofía: {list(sofia_df.columns)}")

    # Filtrar registros de Sofía por Código Ficha
    logger.info(f"Filtrando registros de Sofía por el Código de Ficha: {codigo_ficha}...")
    sofia_filtrado = sofia_df[sofia_df["FICHA"] == int(codigo_ficha)]
    if sofia_filtrado.empty:
        logger.error(f"No se encontraron registros en Sofía para la ficha {codigo_ficha}.")
        raise ValueError(f"No se encontraron registros en Sofía para la ficha {codigo_ficha}.")
    logger.info(f"Registros filtrados en Sofía: {len(sofia_filtrado)}")

    # Función para limpiar y convertir números de documento
    def limpiar_documento(doc):
        # Convertir a cadena, eliminar espacios y caracteres no numéricos
        doc_str = str(doc).replace('E+', '').replace('.', '')
        return ''.join(filter(str.isdigit, doc_str))

    # Crear columnas de comparación en formato de texto
    logger.info("Preparando columnas de comparación...")
    instru_df["DOCUMENTO_DE_IDENTIFICACION_COMP"] = instru_df["DOCUMENTO_DE_IDENTIFICACION_INSTRU"].apply(limpiar_documento)
    sofia_filtrado["DOCUMENTO_DE_IDENTIFICACION_COMP"] = sofia_filtrado["DOCUMENTO_DE_IDENTIFICACION_SOFIA"].apply(limpiar_documento)
    
    instru_df["DOCUMENTO_DE_IDENTIFICACION_INSTRU"] = instru_df["DOCUMENTO_DE_IDENTIFICACION_COMP"]
    sofia_filtrado["DOCUMENTO_DE_IDENTIFICACION_SOFIA"] = sofia_filtrado["DOCUMENTO_DE_IDENTIFICACION_COMP"]

    # Realizar la validación con merge preferencial
    logger.info("Comenzando la validación y comparación de datos...")
    validacion_df = pd.merge(
        instru_df,
        sofia_filtrado,                                                                                                                                                                     
        left_on="DOCUMENTO_DE_IDENTIFICACION_COMP",
        right_on="DOCUMENTO_DE_IDENTIFICACION_COMP",
        how='outer',
        suffixes=('_instru', '_sofia')
    )

    def verificar_discrepancias(row):
        discrepancias = []
        
        # Comparación de todos los datos como texto
        if pd.isna(row['DOCUMENTO_DE_IDENTIFICACION_INSTRU']) and not pd.isna(row['DOCUMENTO_DE_IDENTIFICACION_SOFIA']):
            discrepancias.append("Documento sólo en Sofía")
        elif not pd.isna(row['DOCUMENTO_DE_IDENTIFICACION_INSTRU']) and pd.isna(row['DOCUMENTO_DE_IDENTIFICACION_SOFIA']):
            discrepancias.append("Documento sólo en Instructores")
        else:
            # Comparar Tipo de Documento
            tipo_instru = str(row['TIPO_INSTRU']).strip()
            tipo_sofia = str(row['TIPO_SOFIA']).strip()
            if tipo_instru != tipo_sofia:
                discrepancias.append(f"Discrepancia en Tipo de Documento: Instructores ({tipo_instru}) vs Sofía ({tipo_sofia})")
            
            # Comparar Nombre Completo
            nombre_instru = str(row['NOMBRE_COMPLETO_SENA_INSTRU']).strip()
            nombre_sofia = str(row['NOMBRE_COMPLETO_SENA_SOFIA']).strip()
            if nombre_instru != nombre_sofia:
                discrepancias.append(f"Discrepancia en Nombre: Instructores ({nombre_instru}) vs Sofía ({nombre_sofia})")

        return "FALSO - " + "; ".join(discrepancias) if discrepancias else "VERDADERO"

    # Agregar columna de coincidencia
    validacion_df['COINCIDENCIA'] = validacion_df.apply(verificar_discrepancias, axis=1)

    # Eliminar duplicados basados en el número de documento limpio
    validacion_df.drop_duplicates(subset=['DOCUMENTO_DE_IDENTIFICACION_COMP'], keep='first', inplace=True)

    # Renombrar columnas y reordenando
    logger.info("Renombrando columnas y reordenando...")
    columnas_ordenadas = [
        "FICHA", "TIPO_PROGRAMA", "NIVEL_FORMACION", "DENOMINACION_PROGRAMA",
        "TIPO_INSTRU", "TIPO_SOFIA",
        "DOCUMENTO_DE_IDENTIFICACION_INSTRU", "DOCUMENTO_DE_IDENTIFICACION_SOFIA",
        "NOMBRE_COMPLETO_SENA_INSTRU", "NOMBRE_COMPLETO_SENA_SOFIA", "COINCIDENCIA"
    ]
    validacion_df = validacion_df[columnas_ordenadas]
    validacion_df.rename(columns={
        "FICHA": "Ficha",
        "TIPO_PROGRAMA": "Tipo Programa",
        "NIVEL_FORMACION": "Nivel Formación",
        "DENOMINACION_PROGRAMA": "Denominación Programa",
        "TIPO_INSTRU": "Tipo Documento Instructores",
        "TIPO_SOFIA": "Tipo Documento Sofía",
        "DOCUMENTO_DE_IDENTIFICACION_INSTRU": "No. Documento Instructores",
        "DOCUMENTO_DE_IDENTIFICACION_SOFIA": "No. Documento Sofía",
        "NOMBRE_COMPLETO_SENA_INSTRU": "Nombre Aprendiz Instructores",
        "NOMBRE_COMPLETO_SENA_SOFIA": "Nombre Aprendiz Sofía",
        "COINCIDENCIA": "COINCIDENCIA"
    }, inplace=True)

    output_path = "resultado_validacion.xlsx"
    validacion_df.to_excel(output_path, index=False)
    ajustar_ancho_columnas(output_path)
    logger.info(f"Archivo de resultado guardado en: {output_path}")

    return open(output_path, "rb")