import openpyxl 
from openpyxl.styles import PatternFill, Font, Alignment
from ComponentesBackEnd.logger_configuracion import logger



def ajustar_ancho_columnas(archivo):
    wb = openpyxl.load_workbook(archivo)
    hoja = wb.active

    for columna in hoja.columns:
        max_length = 0
        columna_letra = columna[0].column_letter
        for celda in columna:
            try:
                if celda.value:
                    max_length = max(max_length, len(str(celda.value)))
            except Exception as e:
                logger.warning(f"Error ajustando la celda {celda.coordinate}: {e}")

        hoja.column_dimensions[columna_letra].width = max_length + 2

    logger.info("Aplicando formato a los encabezados...")
    encabezado_fila = hoja[1]
    estilo_encabezado = {
        "fill": PatternFill(start_color="C6D8FD", end_color="C6D8FD", fill_type="solid"),
        "font": Font(size=13, bold=True, color="000000"),
        "alignment": Alignment(horizontal="center", vertical="center"),
    }

    for celda in encabezado_fila:
        celda.fill = estilo_encabezado["fill"]
        celda.font = estilo_encabezado["font"]
        celda.alignment = estilo_encabezado["alignment"]
    
    
    logger.info("Aplicando Formato al cuerpo del archivo")
    cuerpo_archivo = hoja[1]
    estilo_cuerpo = {
        "font": Font(size=12, color="000000"),
        "alignment": Alignment(horizontal="left", vertical="center"),
    }
    
    for celda in cuerpo_archivo:
        celda.font = estilo_cuerpo["font"]
        celda.alignment = estilo_cuerpo["alignment"]
        
        

    # Encontrar la columna de Juicios Evaluativos
    juicios_col = None
    for col in range(1, hoja.max_column + 1):
        if hoja.cell(row=1, column=col).value == "Juicios Evaluativos":
            juicios_col = col
            break

    if juicios_col:
        for row in range(2, hoja.max_row + 1):
            valor = hoja.cell(row=row, column=juicios_col).value  # Obtén el valor de la celda
            try:
                if valor and "Aprobó" in str(valor):  # Verifica si tiene un formato esperado
                    # Intenta extraer los números de "Aprobó X de Y juicios"
                    partes = str(valor).replace("Aprobó", "").replace("juicios", "").strip().split("de")
                    aprobados, total = map(int, partes)  # Convierte los números
                    if aprobados == total:
                        color = "90EE90"  # Verde claro si aprobados == total
                    else:
                        color = "FFFFE0"  # Amarillo opaco si no coincide
                else:
                    color = "FFFFE0"  # Amarillo opaco si el valor no cumple con el formato esperado
            except (ValueError, AttributeError):
                # Manejo de excepciones si el formato no es válido
                color = "FFFFE0"  # Amarillo opaco como predeterminado para errores

            # Aplica el color a la celda
            hoja.cell(row=row, column=juicios_col).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        logger.info("Columna de Juicios Evaluativos coloreada según criterios.")

                                                                          
                

    wb.save(archivo)
    logger.info(f"Ancho de columnas ajustado y formato aplicado en el archivo: {archivo}")